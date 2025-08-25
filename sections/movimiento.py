from PyQt6.QtWidgets import QWidget, QTableWidgetItem, QMessageBox, QFileDialog
from PyQt6.QtCore import QDate
from ui_files.movimiento_ui import Ui_Form
import pandas as pd
import datetime
import locale
from openpyxl import load_workbook
import re


class MovimientoWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        # Set locale for current date
        locale.setlocale(locale.LC_TIME, "Spanish_Mexico.1252")

        # Set ReadOnly the registrar button
        self.ui.registrar_btn.setDisabled(True)

        # Set current date
        self.ui.aplicacion_input.setDate(datetime.datetime.now())

        # Set current termino date
        self.calculate_termino_date()

        # Calculate new fecha_termino
        self.ui.aplicacion_input.dateChanged.connect(self.calculate_termino_date)

        # Set Centro Gestor
        self.ui.centro_input.setText("SECRETARÍA DE SEGURIDAD CIUDADANA")

        # Load recent records when the widget is initialized
        self.load_recent_records()

        # Load employee data if there is a value in the no_emp_input
        self.ui.no_emp_input.returnPressed.connect(self.load_employee_data)

        # Connect registrar button
        self.ui.registrar_btn.clicked.connect(self.create_excel_document)

    def load_employee_data(self):
        # Load employee data from Excel
        try:
            df = self.read_main_database()
        except Exception:
            QMessageBox.warning(self, "Leer Base de Datos", "Error al leer la Base de Datos de Empleados.")
            return
        
        no_empleado = self.ui.no_emp_input.text()

        if no_empleado.isdigit():
            no_empleado = int(no_empleado)
            if no_empleado in df['No. DE EMPLEADO'].values:
                try:
                    row = df[df['No. DE EMPLEADO'] == no_empleado].index[0]
                    self.ui.nombre_input.setText(df.loc[row, 'APELLIDO PATERNO'] + ' ' + df.loc[row, 'APELLIDO MATERNO'] + ' ' + df.loc[row, 'NOMBRE'])
                    self.ui.curp_input.setText(df.loc[row, 'C.U.R.P.'])
                    self.ui.rfc_input.setText(df.loc[row, 'R.F.C.'])
                    self.ui.ingreso_input.setDate(df.loc[row, 'INGRESO'])
                    self.ui.direccion_input.setText(df.loc[row, 'Adscripción'])
                    self.ui.categoria_input.setText(df.loc[row, 'GRADO'])
                    self.ui.clave_input.setText(df.loc[row, 'CLAVE'])
                    self.ui.nivel_input.setText(df.loc[row, 'CLAVE'])
                    
                    self.ui.registrar_btn.setDisabled(False)
                except Exception as e:
                    QMessageBox.warning(self, "Error al obtener datos", f'Asegúrate de que el nombre de las columnas sean correctas. Error: {e}')
                    self.ui.registrar_btn.setDisabled(True)
                    self.clear_inputs()
            else:
                QMessageBox.warning(self, "Empleado no encontrado", f'El empleado {no_empleado} no se encuentra en la base de datos.')
                self.ui.registrar_btn.setDisabled(True)
                self.ui.no_emp_input.setText(str(no_empleado))
                self.clear_inputs()

    def get_input_data(self):
        data = {
            "No.Empleado": self.ui.no_emp_input.text(),
            "Nombre": self.ui.nombre_input.text(),
            "CURP": self.ui.curp_input.text(),
            "RFC": self.ui.rfc_input.text(),
            "Centro Gestor": self.ui.centro_input.text(),
            "Dirección": self.ui.direccion_input.text(),
            "Categoría": self.ui.categoria_input.text(),
            "Clave": self.ui.clave_input.text(),
            "Nivel": self.ui.nivel_input.text(),
            "Fecha de Inicio": self.ui.ingreso_input.text(),
            "Fecha de Termino": self.ui.termino_input.text(),
            "Fecha de aplicación": self.ui.aplicacion_input.text(),
            "Tipo": self.ui.tipo_input.currentText(),
            "Vigencia": self.ui.vigencia_input.currentText(),
            "Plaza": self.ui.plaza_input.currentText(),
        }

        # Get type of movimiento
        TNI, TC, TB, TP, TR, TO = self.def_type_of_movimiento(data['Tipo'])
        data['TNI'], data['TC'], data['TB'], data['TP'], data['TR'], data['TO'] = TNI, TC, TB, TP, TR, TO

        # Get vigencia
        V, N = self.def_vigencia(data['Vigencia'])
        data['VD'], data['VPTF'] = V, N

        # Get plaza
        PC, PB, PE, PP = self.def_plaza(data['Plaza'])
        data['PC'], data['PB'], data['PE'], data['PP'] = PC, PB, PE, PP

        # Get date for day, month and year
        # Inicio
        day, month, year = self.get_day_month_year(data['Fecha de Inicio'])
        data['DI'], data['MI'], data['AI'] = day, month, year
        # Termino
        day, month, year = self.get_day_month_year(data['Fecha de Termino'])
        data['DT'], data['MT'], data['AT'] = day, month, year
        return data
    
    def def_type_of_movimiento(self, tipo):
        TNI, TC, TB, TP, TR, TO = '', '', '', '', '', ''
        if tipo == 'Nuevo ingreso':
            TNI = 'X'
        elif tipo == 'Cambio':
            TC = 'X'
        elif tipo == 'Baja':
            TB = 'X'
        elif tipo == 'Promoción':
            TP = 'X'
        elif tipo == 'Reinstalación':
            TR = 'X'
        elif tipo == 'Otro':
            TO = 'X'
        return TNI, TC, TB, TP, TR, TO
    
    def def_vigencia(self, vigencia):
        VD, VPTF = '', ''
        if vigencia == 'Definitivo':
            VD = 'X'
        elif vigencia == 'Por tiempo fijo':
            VPTF = 'X'
        return VD, VPTF
    
    def def_plaza(self, plaza):
        PC, PB, PE, PP = '', '', '', ''
        if plaza == 'Confianza':
            PC = 'X'
        elif plaza == 'Base':
            PB = 'X'
        elif plaza == 'Eventual':
            PE = 'X'
        elif plaza == 'Provisional':
            PP = 'X'
        return PC, PB, PE, PP

    def calculate_termino_date(self):
        try:
            fecha_aplicacion = datetime.datetime.strptime(self.ui.aplicacion_input.text(), '%d/%m/%Y')
            termino = fecha_aplicacion - datetime.timedelta(days=1)

            # Convert to QDate
            termino_qdate = QDate(termino.year, termino.month, termino.day)

            self.ui.termino_input.setDate(termino_qdate)
        except Exception as e:
            print('Error al calcular la fecha de termino. Error:', e)

    def get_day_month_year(self, date):
        date = datetime.datetime.strptime(date, '%d/%m/%Y')
        day = date.day
        month = date.month
        year = date.year
        return day, month, year
    
    def register_data_on_database(self):
        # Get data from inputs
        data = self.get_input_data()

        # Delete the date values
        data.pop('DI')
        data.pop('MI')
        data.pop('AI')
        data.pop('DT')
        data.pop('MT')
        data.pop('AT')

        # Drop the extra values
        data.pop('TNI')
        data.pop('TC')
        data.pop('TB')
        data.pop('TP')
        data.pop('TR')
        data.pop('TO')
        data.pop('V')
        data.pop('N')
        data.pop('PC')
        data.pop('PB')
        data.pop('PE')

        # Create a new DataFrame with the data
        nuevo = pd.DataFrame([data])
        try:
            df = pd.read_excel("./data/database/movimiento_personal.xlsx") if pd.io.common.file_exists("./data/database/movimiento_personal.xlsx") else pd.DataFrame()
        except Exception:
            df = pd.DataFrame()

        # Append the new DataFrame to the existing DataFrame
        df = pd.concat([df, nuevo], ignore_index=True)
        df.to_excel("./data/database/movimiento_personal.xlsx", index=False)

        self.clear_inputs()
        self.load_recent_records()
    
    def create_excel_document(self):
        try:
            # Cargar plantilla Excel
            template_path = './data/templates/movimiento_personal.xlsx'
            wb = load_workbook(template_path)
            ws = wb.active
            
            data = self.get_input_data()

            # Convertir y formatear los datos
            for key in data:
                try:
                    if key != "Categoría" or key != "Nombre":
                        data[key] = str(data[key]).upper()
                except Exception:
                    pass

            # Función para verificar si una celda es parte de un rango combinado
            def is_merged(cell):
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        return merged_range.start_cell  # Devuelve la celda principal
                return None

            # Función para reemplazar marcadores
            def replace_markers_in_cell(cell_value, replacements):
                if cell_value and isinstance(cell_value, str):
                    for key, value in replacements.items():
                        pattern = re.compile(r'\{\{\s*' + re.escape(key) + r'\s*\}\}', re.IGNORECASE)
                        cell_value = pattern.sub(str(value), cell_value)
                return cell_value

            # Buscar y reemplazar en todas las celdas
            for row in ws.iter_rows():
                for cell in row:
                    # Verificar si la celda es parte de un rango combinado
                    main_cell = is_merged(cell)
                    target_cell = main_cell if main_cell else cell
                    
                    # Solo procesar si es la celda principal o no está combinada
                    if target_cell == cell:
                        # Reemplazar en el valor de la celda
                        target_cell.value = replace_markers_in_cell(target_cell.value, data)
                        
                        # Reemplazar en comentarios si los hay
                        if target_cell.comment:
                            comment_text = replace_markers_in_cell(target_cell.comment.text, data)
                            target_cell.comment.text = comment_text
            
            # Nombre sugerido para el archivo
            suggested_name = f"{data['Nombre']}_MOVIMIENTO_PERSONAL.xlsx"
            
            # Preguntar al usuario dónde guardar
            save_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Documento",
                suggested_name,
                "Excel Files (*.xlsx)"
            )
            
            if save_path:
                wb.save(save_path)
                self.register_data_on_database()
                QMessageBox.information(self, "Guardado", f"Documento guardado en:\n{save_path}")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Ocurrió un error al generar el documento:\n{str(e)}")

    def clear_inputs(self):
        self.ui.no_emp_input.clear()
        self.ui.nombre_input.clear()
        self.ui.curp_input.clear()
        self.ui.rfc_input.clear()
        self.ui.direccion_input.clear()
        self.ui.categoria_input.clear()
        self.ui.clave_input.clear()
        self.ui.nivel_input.clear()
        self.ui.ingreso_input.clear()
        self.ui.termino_input.clear()
        self.ui.aplicacion_input.clear()
        self.ui.centro_input.setText("SECRETARÍA DE SEGURIDAD CIUDADANA")
    
    def read_main_database(self):
        try:
            df = pd.read_excel("./data/database/main.xlsx")
        except Exception as e:
            QMessageBox.warning(self, "Leer Base de Datos", f"Error al leer la Base de Datos de Empleados: {str(e)}")
            return

        return df

    def load_recent_records(self):
        # Load recent records from quinquenio database
        try:
            df = pd.read_excel("./data/database/modif_vacaciones.xlsx")
        except Exception:
            self.ui.table_display.setRowCount(0)
            self.ui.table_display.setColumnCount(0)
            return

        self.ui.table_display.setRowCount(len(df))
        self.ui.table_display.setColumnCount(len(df.columns))
        self.ui.table_display.setHorizontalHeaderLabels(df.columns)

        # Reverse the DataFrame (most recent records first)
        df = df.iloc[::-1].reset_index(drop=True)

        # Set the table data
        for i in range(len(df)):
            for j in range(len(df.columns)):
                item = QTableWidgetItem(str(df.iat[i, j]))
                self.ui.table_display.setItem(i, j, item)