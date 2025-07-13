from PyQt6.QtWidgets import QWidget, QTableWidgetItem, QMessageBox, QFileDialog
from ui_files.tiempoxtiempo_ui import Ui_Form
import pandas as pd
from openpyxl import load_workbook
import re

class TiempoXTiempoWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        # Set ReadOnly the registrar button
        self.ui.registrar_btn.setDisabled(True)

        # Load recent records when the widget is initialized
        self.load_recent_records()

        # Load employee data if there is a value in the no_emp_input
        self.ui.no_emp_input.returnPressed.connect(self.load_employee_data)

        # Connect registrar button
        self.ui.registrar_btn.clicked.connect(self.create_excel_document)

    def load_employee_data(self):
        # Load employee data from Excel
        try:
            df = self.read_main_database()
        except Exception:
            QMessageBox.warning(self, "Leer Base de Datos", "Error al leer la Base de Datos de Empleados.")
            return
        
        no_empleado = self.ui.no_emp_input.text()

        if no_empleado.isdigit():
            no_empleado = int(no_empleado)
            if no_empleado in df['No. DE EMPLEADO'].values:
                try:
                    row = df[df['No. DE EMPLEADO'] == no_empleado].index[0]
                    self.ui.nombre_input.setText(df.loc[row, 'NOMBRE COMPLETO'])
                    self.ui.area_input.setText(df.loc[row, 'Adscripción'])
                    self.ui.registrar_btn.setDisabled(False)
                except Exception as e:
                    QMessageBox.warning(self, "Error al obtener datos", f'Asegúrate de que el nombre de las columnas sean correctas. Error: {e}')
                    self.ui.registrar_btn.setDisabled(True)
                    self.clear_inputs()
            else:
                QMessageBox.warning(self, "Empleado no encontrado", f'El empleado {no_empleado} no se encuentra en la base de datos.')
                self.ui.registrar_btn.setDisabled(True)
                self.ui.no_emp_input.setText(str(no_empleado))
                self.clear_inputs()

    def get_input_data(self):
        data = {
            "No.Empleado": self.ui.no_emp_input.text(),
            "Nombre": self.ui.nombre_input.text(),
            "CEGE": self.ui.cege_input.text(),
            "Área": self.ui.area_input.text(),
            "Fecha de Tiempo Extra": self.ui.fecha_tiempo_input.text(),
            "Motivo": self.ui.motivo_input.text(),
            "Horas Extra": self.ui.horas_extra_input.text() + " HORAS",
            "Día Otorgado": self.ui.fecha_otorgado_input.text()
        }

        return data
    
    def register_data_on_database(self):
        # Get data from inputs
        data = self.get_input_data()

        # Create a new DataFrame with the data
        nuevo = pd.DataFrame([data])
        try:
            df = pd.read_excel("./data/database/tiempoxtiempo.xlsx") if pd.io.common.file_exists("./data/database/tiempoxtiempo.xlsx") else pd.DataFrame()
        except Exception:
            df = pd.DataFrame()

        # Append the new DataFrame to the existing DataFrame
        df = pd.concat([df, nuevo], ignore_index=True)
        df.to_excel("./data/database/tiempoxtiempo.xlsx", index=False)

        self.clear_inputs()
        self.load_recent_records()
    
    def create_excel_document(self):
        try:
            # Cargar plantilla Excel
            template_path = './data/templates/tiempoxtiempo.xlsx'
            wb = load_workbook(template_path)
            ws = wb.active
            
            data = self.get_input_data()

            # Convertir y formatear los datos
            for key in data:
                try:
                    data[key] = str(data[key]).upper()
                except Exception:
                    pass

            # Función para verificar si una celda es parte de un rango combinado
            def is_merged(cell):
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        return merged_range.start_cell  # Devuelve la celda principal
                return None

            # Función para reemplazar marcadores
            def replace_markers_in_cell(cell_value, replacements):
                if cell_value and isinstance(cell_value, str):
                    for key, value in replacements.items():
                        pattern = re.compile(r'\{\{\s*' + re.escape(key) + r'\s*\}\}', re.IGNORECASE)
                        cell_value = pattern.sub(str(value), cell_value)
                return cell_value

            # Buscar y reemplazar en todas las celdas
            for row in ws.iter_rows():
                for cell in row:
                    # Verificar si la celda es parte de un rango combinado
                    main_cell = is_merged(cell)
                    target_cell = main_cell if main_cell else cell
                    
                    # Solo procesar si es la celda principal o no está combinada
                    if target_cell == cell:
                        # Reemplazar en el valor de la celda
                        target_cell.value = replace_markers_in_cell(target_cell.value, data)
                        
                        # Reemplazar en comentarios si los hay
                        if target_cell.comment:
                            comment_text = replace_markers_in_cell(target_cell.comment.text, data)
                            target_cell.comment.text = comment_text
            
            # Nombre sugerido para el archivo
            nombre_parts = data.get("Nombre", "").split()
            suggested_name = f"{nombre_parts[-1] if nombre_parts else ''}_{data.get('No.Empleado', '')}_TIEMPO_POR_TIEMPO.xlsx"
            
            # Preguntar al usuario dónde guardar
            save_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Documento",
                suggested_name,
                "Excel Files (*.xlsx)"
            )
            
            if save_path:
                wb.save(save_path)
                self.register_data_on_database()
                QMessageBox.information(self, "Guardado", f"Documento guardado en:\n{save_path}")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Ocurrió un error al generar el documento:\n{str(e)}")

    def clear_inputs(self):
        self.ui.no_emp_input.clear()
        self.ui.nombre_input.clear()
        self.ui.area_input.clear()
        self.ui.registrar_btn.setDisabled(True)
    
    def read_main_database(self):
        try:
            df = pd.read_excel("./data/database/main.xlsx")
        except Exception as e:
            QMessageBox.warning(self, "Leer Base de Datos", f"Error al leer la Base de Datos de Empleados: {str(e)}")
            return

        return df

    def load_recent_records(self):
        # Load recent records from quinquenio database
        try:
            df = pd.read_excel("./data/database/tiempoxtiempo.xlsx")
        except Exception:
            self.ui.table_display.setRowCount(0)
            self.ui.table_display.setColumnCount(0)
            return

        self.ui.table_display.setRowCount(len(df))
        self.ui.table_display.setColumnCount(len(df.columns))
        self.ui.table_display.setHorizontalHeaderLabels(df.columns)

        # Reverse the DataFrame (most recent records first)
        df = df.iloc[::-1].reset_index(drop=True)

        # Set the table data
        for i in range(len(df)):
            for j in range(len(df.columns)):
                item = QTableWidgetItem(str(df.iat[i, j]))
                self.ui.table_display.setItem(i, j, item)